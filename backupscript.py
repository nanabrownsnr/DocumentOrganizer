
import requests
import json
import streamlit as st
import os
import re
# from pdf2image import convert_from_bytes
import io
import pdftotext
import openpyxl

st.set_page_config(page_title="Demo", page_icon="📈")

# Convert PDF to Image
# def pdf_to_image(uploaded_file):
#     pdf_images = convert_from_bytes(uploaded_file.read())
#     return pdf_images

def Document_Classification(input_file):
    files = [('pdf_file', (input_file.name, input_file.getvalue(), 'application/pdf'))]

    response = requests.post('https://text-document-classification-vla-1.ai-rest-services.ch/api/v1/classify/', files=files)

    return response.json()

# def Text_Classification(input_file):
#     files = [('pdf_file', (input_file.name, input_file.getvalue(), 'application/pdf'))]

#     response = requests.post('https://text-document-classification-vla-1.ai-rest-services.ch/api/v1/classify/', files=files)

#     return response.json()

def Text_Classification(text):
    review_categories="Patient File, STI Test Form, MRI Test Results"
    response = requests.post('https://text-classification-disbart-mnli-12-1.ai-rest-services.ch/api/v1/classify',
                             data=json.dumps({"text": text, "labels": review_categories}))

    return response.json()["answer"]["labels"][0]

def call_api_image(image, question):
    img_byte_arr = io.BytesIO()
    image.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()
    data = {"questions": question}
    response = requests.post(url="https://image-question-answer-layoutlm-v2.ai-rest-services.ch/api/v1/answer",
                             files=[('file', ('file', img_byte_arr, 'image/png'))], data=data)
    if response.status_code == 200:
        return response.json()[0]["answer"]
    else:
        st.write("API request failed with status code:", response.status_code)
        return None

def call_api_pdf(document, question):
    data = {"questions": question}
    response = requests.post(url="https://text-document-question-answer-bert-large.ai-rest-services.ch/api/v1/answer",
                             files=[('pdf_file', (document.name, document, 'application/pdf'))], data=data)
    if response.status_code == 200:
        return response.json()[0]["answer"]
    else:
        st.write("API request failed with status code:", response.status_code)
        return None


def call_api_text(text, question):
    data = {
        "question": question,
        "context": text
    }
    response = requests.post(url="https://text-question-answer-xlm-roberta-squad2.ai-rest-services.ch/question",
                             data=json.dumps(data))
    if response.status_code == 200:
        return response.json()["answer"]
    else:
        st.write("API request failed with status code:", response.status_code)
        return None


def sort_file(category, file, output_folder):
    success = False
    try:
        if category is not None:
            category = re.sub('[^A-Za-z0-9]+', '', category)

            if output_folder[len(output_folder) - 1] != '/':
                output_folder = output_folder + '/'

            newpath = f"{output_folder}{category}"
            if not os.path.exists(newpath):
                os.mkdir(newpath)

            with open(f"{newpath}/{file.name}", "bw") as new_file:
                new_file.write(file.getvalue())
                success = True
    except Exception:
        return False
    else:
        return success
    
def update_excel_file(filename, patientname,filecategory):
    success = False
    try:
        if not os.path.exists('patient_info.xlsx'):
            workbook = openpyxl.Workbook()
            sheet = workbook.create_sheet('Sheet1')
            sheet['A1'] = "Patient Name"
            sheet['B1'] = "Filename"
            sheet['C1'] = "File Type"
        else:
            workbook = openpyxl.load_workbook('patient_info.xlsx')
            
        sheet = workbook['Sheet1']
        row = sheet.max_row + 1
        sheet['A' + str(row)] = patientname
        sheet['B' + str(row)] = filename
        sheet['C' + str(row)] = filecategory

        workbook.save('patient_info.xlsx')
        success = True
    except Exception:
        return False
    else:
        return success

def extract_data(document):
    document_text = []
    with open(document, "rb") as f:
        pdf = pdftotext.PDF(f)

        for page in pdf:
            document_text.append(page)
    data = " ".join(map(str, document_text))
    return data
def extract_data_st(document):
    # pdf_bytes = document.read()
    pdf = pdftotext.PDF(document)
    pdf_content = "\n".join(pdf)
    # with pdftotext.PDF(io.BytesIO(pdf_bytes)) as pdf:
    return pdf_content

st.title("Document Organizer")

files = st.file_uploader("Choose a file", type=['pdf'], accept_multiple_files=True)
question = "What is the patients name?"

if files is not None and st.button("Organize Documents"):
    progress_text = "Operation in progress. Please wait."
    
    my_bar = st.progress(0, text=progress_text)
    for x in range(len(files)):
        file=files[x]
        filename = file.name
        # with st.spinner("Processing"):
        #     # name = call_api_pdf(file, question)
        #     # image = pdf_to_image(file)[0]
        #     # st.image(image)
        #     # name=call_api_image(image, question)
        #     # text = extract_data_st(file)
        text= Document_Classification(file)["document_text"]
        # st.write(text)
        name = call_api_text(text, question)
        category = Text_Classification(text)

        if name is not None:
            # st.write(f"Patient Name: {name}")
            sort_file(name, file, ".")
            #     st.write("File sorted")
            # else:
            #     st.write("Sorting failed")

            update_excel_file(filename,name,category)
            #     st.write("Database Updated")
            # else:
            #     st.write("Update failed")

        else:
            st.write("Unable to identify Patient Name")
        
        my_bar.progress(int(100 * (x+1) / len(files)), text=progress_text)

    st.success('Files Sorted.')
if not files:
    st.warning("Upload a file")
